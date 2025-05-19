import os
import re
import openpyxl
from statistics import mean

duration_pattern = re.compile(r'Duration:\s+([0-9.]+)\s+ms')


def extract_durations(folder_path):
    parallel_durations = {}
    sequential_durations = {}

    for filename in sorted(os.listdir(folder_path)):
        if filename.endswith('.txt'):
            filepath = os.path.join(folder_path, filename)
            match = re.search(r'(\d+)', filename)
            if match:
                id_ = int(match.group(1))
                with open(filepath, 'r') as file:
                    content = file.read()
                    match_duration = duration_pattern.search(content)
                    if match_duration:
                        duration = float(match_duration.group(1))
                        if filename.startswith('Parallel'):
                            parallel_durations[id_] = duration
                        elif filename.startswith('Unparallel'):
                            sequential_durations[id_] = duration
                        elif filename.startswith('Sequential'):
                            sequential_durations[id_] = duration

    return parallel_durations, sequential_durations


def write_to_excel(parallel_durations, sequential_durations, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Headers
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=2, column=2, value="Parallel Duration (ms)")
    ws.cell(row=2, column=3, value="Sequential Duration (ms)")

    # Merge header cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)

    # Fill table
    all_ids = sorted(set(parallel_durations.keys()) | set(sequential_durations.keys()))
    for idx, id_ in enumerate(all_ids, start=3):
        ws.cell(row=idx, column=1, value=id_)
        ws.cell(row=idx, column=2, value=parallel_durations.get(id_))
        ws.cell(row=idx, column=3, value=sequential_durations.get(id_))

    # Means
    last_row = len(all_ids) + 3
    ws.cell(row=last_row, column=1, value="Mean")
    ws.cell(row=last_row, column=2, value=mean(parallel_durations.values()))
    ws.cell(row=last_row, column=3, value=mean(sequential_durations.values()))

    # Save workbook
    wb.save(output_file)
    print(f"Excel report created: {output_file}")


# === Client Logs ===
client_folder_path = 'Utility/Clog-1/'
client_parallel, client_seq = extract_durations(client_folder_path)
write_to_excel(client_parallel, client_seq, 'comparison_table_ratio_0.0_bs_1000_client.xlsx')

# === Server Logs ===
server_folder_path = 'Server_RMI/Slog-1/'
server_parallel, server_seq = extract_durations(server_folder_path)
write_to_excel(server_parallel, server_seq, 'comparison_table_ratio_0.0_bs_1000_server.xlsx')

print("Excel report created.")

# Rename folders (optional cleanup/archive step)
os.rename('Utility/Clog-1', 'Utility/Clog-1-r-0.0-b-1000')
os.rename('Server_RMI/Slog-1', 'Server_RMI/Slog-1-r-0.0-b-1000')
print("Folders renamed successfully.")
